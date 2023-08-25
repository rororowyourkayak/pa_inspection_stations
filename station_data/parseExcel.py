def extract(ws):
    counter = 1
    sheetData = []
    for sheet in ws:
        if counter % 2 == 0:
            for row in sheet.iter_rows(min_row=3):
                rowData = []
                for cell in row:
                    rowData.append(cell.value)
                sheetData.append(rowData)      
               
        counter+=1
       
    return sheetData

import openpyxl as xl

finalDataWB = xl.Workbook()
finalData = finalDataWB.active
finalData.append(['county','station_name','ois_number','station_address','mailing_address','phone_number','passenger_cars_and_light_trucks','medium_trucks','heavy_trucks','motorcycle','trailer_less_10000','trailer_greater_10000'])

wb1 = xl.load_workbook('stations1199.xlsx')
wb2 = xl.load_workbook('stations200399.xlsx')
wb3 = xl.load_workbook('stations400599.xlsx')
wb4 = xl.load_workbook('stations600604.xlsx')

for row in extract(wb1):
    finalData.append(row)
for row in extract(wb2):
    finalData.append(row)
for row in extract(wb3):
    finalData.append(row)
for row in extract(wb4):
    finalData.append(row)

finalDataWB.save('stations.xlsx')
quit()